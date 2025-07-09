#!/usr/bin/env python3
import sys
import os
import datetime
from openpyxl import load_workbook

def get_now_time_str():
    now_time = datetime.datetime.now()
    return 'File created time : ' + now_time.strftime("%Y-%m-%d %H:%M:%S")

def collect_variable_list(input_file, sheet_name):
    workbook = load_workbook(filename=input_file, data_only=True)
    worksheet = workbook[sheet_name]
    variable_list = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data_type = row[1]
        variable_name = row[2]
        comment = row[3]
        if variable_name and data_type:
            variable_list.append({
                'variable_name': variable_name,
                'data_type': data_type,
                'comment': comment
            })
    return variable_list

def generate_h_file(variable_list, output_hfile, sheet_name, tool_name, time_str, input_file):
    with open(output_hfile, 'w') as f:
        print('Generating .h file...')

        header = (
            "/************************************************************************************************\n"
            f"{time_str}\n"
            f"Tool Name :         {tool_name} + {input_file}\n"
            f"Filename:           {output_hfile}\n"
            "Author:             Ai\n"
            "Description:        Generate VFB interface\n"
            "/************************************************************************************************/\n"
            "/************************************************************************************************\n"
            "                                   C O P Y R I G H T  \n"
            "-------------------------------------------------------------------------------  \n"
            "Copyright (c) 2013-2025 by Shenzhen Southern Dare Automotive Electronics Co.,Ltd.. All rights reserved.  \n"
            "/************************************************************************************************/\n\n"
            f"#ifndef _{sheet_name.upper()}_H_\n"
            f"#define _{sheet_name.upper()}_H_\n\n"
            "#include \"Common.h\"\n\n\n"
            "/*********************************Function declaration*********************************************/\n"
        )
        f.write(header)

        for item in variable_list:
            f.write(f"extern void Rte_SetVfb_{item['variable_name']}({item['data_type']} SetValue);\n")
            f.write(f"extern {item['data_type']} Rte_GetVfb_{item['variable_name']}(void);\n\n")

        footer = (
            "/***************************end****************************/\n\n\n"
            "#endif   /* _RTE_VFBINTERFACE_H_ */\n"
        )
        f.write(footer)

def generate_c_file(variable_list, output_cfile, output_hfile, tool_name, time_str, input_file):
    with open(output_cfile, 'w') as f:
        print('Generating .c file...')

        header = (
            "/************************************************************************************************\n"
            f"{time_str}\n"
            f"Tool Name :         {tool_name} + {input_file}\n"
            f"Filename:           {output_cfile}\n"
            "Author:             Ai\n"
            "Description:        Generate VFB interface\n"
            "/************************************************************************************************/\n"
            "/************************************************************************************************\n"
            "                                   C O P Y R I G H T  \n"
            "-------------------------------------------------------------------------------  \n"
            "Copyright (c) 2013-2025 by Shenzhen Southern Dare Automotive Electronics Co.,Ltd.. All rights reserved.  \n"
            "/************************************************************************************************/\n\n"
            f"#include \"{output_hfile}\"\n\n"
            "/*********************************Variable Definition *******************************************/\n"
        )
        f.write(header)

        for item in variable_list:
            f.write(f"#if 0\n{item['comment']}\n#endif\n")
            f.write(f"{item['data_type']} {item['variable_name']} = 0;\n\n")

        f.write("\n/*********************************Get/Set Function *********************************************/\n")

        for item in variable_list:
            f.write(f"void Rte_SetVfb_{item['variable_name']}({item['data_type']} SetValue)\n{{\n    {item['variable_name']} = SetValue;\n}}\n")
            f.write(f"{item['data_type']} Rte_GetVfb_{item['variable_name']}(void)\n{{\n    return {item['variable_name']};\n}}\n\n")

        f.write("/***************************end****************************/\n")


def generate_rte_files(input_file, sheet_name):
    tool_name = os.path.basename(__file__)
    time_str = get_now_time_str()

    if not input_file.endswith('.xlsx'):
        print(f"Error: {input_file} is not a .xlsx file! Please convert it to .xlsx format.")
        sys.exit(1)

    output_cfile = f"{sheet_name}.c"
    output_hfile = f"{sheet_name}.h"

    variable_list = collect_variable_list(input_file, sheet_name)

    generate_h_file(variable_list, output_hfile, sheet_name, tool_name, time_str, input_file)
    generate_c_file(variable_list, output_cfile, output_hfile, tool_name, time_str, input_file)

    print('Generation Completed Successfully!')


def main():
    input_file = 'DSM_If_Generate.xlsx'
    # sheet_name = 'RteInterface'
    generate_rte_files(input_file, 'Rte_VfbInterface')
    generate_rte_files(input_file, 'Debug_VfbInterface')


if __name__ == "__main__":
    main()
