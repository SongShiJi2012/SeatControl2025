import openpyxl
# from fun_excel import *
import datetime

NowTime = datetime.datetime.now()
StrNowTime = 'File created time : ' + NowTime.strftime("%Y-%m-%d %H:%M:%S")
print(StrNowTime)

# 配置参数
input_excel = "E2Design.xlsx"  # 输入的Excel文件
sheet_name0 = "E2DataBlockAllocation"  # 数据块地址分配页工作表名称
sheet_name1 = "E2DataDefinition"  # 自定义存储项工作表名称
OutputCFile = 'Srvl_E2Cfg_Generation.c'
OutputHFile = 'Srvl_E2Cfg_Generation.h'

FaultFlag = 0;
# 数据块配置相关-全局变量
DataBlockAllocation_StarAddr = []  # 起始地址
DataBlockAllocation_Size = []  # 长度
DataBlockAllocation_BackupID = []  # 备份ID
DataBlockNumber = 0  # 数据块总数

# 自定义存储项相关-全局变量
Definition_EnumName = []  # 枚举定义
Definition_Dscribe = []  # E2中文描述定义项
Definition_StartIndex = []  # 在数据块中的起始字节下标数
Definition_Size = []  # 长度
Definition_BlockID = []  # 数据块ID
Definition_UseID = []  # 统计使用的ID
Definition_UseID_NO_BACK = []  # 统计使用的ID

def Read_E2DataBlockAllocation(input_file, sheet_name):
    """
    参数:
        input_file: 输入的Excel文件路径
        sheet_name: 要读取的工作表名称
    """
    # 修改全局变量需要声明global
    global FaultFlag
    global DataBlockAllocation_StarAddr
    global DataBlockAllocation_Size
    global DataBlockAllocation_BackupID
    global DataBlockNumber

    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(input_file, data_only=True)

        # 获取指定工作表
        sheet = workbook[sheet_name]

        # 清空全局变量
        DataBlockAllocation_StarAddr = []
        DataBlockAllocation_Size = []
        DataBlockAllocation_BackupID = []

        # 遍历工作表中的所有行

        for i, row in enumerate(sheet.iter_rows(values_only=True), start=2):  # 行号从1开始
            val_0 = sheet.cell(row=i, column=2).value  # 起始地址
            if val_0 is not None:
                val_0 = int(val_0, 16)  # 16进制转换为10进制
                val_1 = sheet.cell(row=i, column=3).value  # 长度
                val_2 = sheet.cell(row=i, column=4).value  # 备份ID
                DataBlockAllocation_StarAddr.append(val_0)
                DataBlockAllocation_Size.append(val_1)
                DataBlockAllocation_BackupID.append(val_2)

                val_3 = sheet.cell(row=i + 1, column=2).value  # 检测长度是否正确
                if val_3 is not None:
                    val_3 = int(val_3, 16)  # 检测长度是否正确
                    if (val_0 + val_1) > val_3:
                        print(f"数据块配置与下一个配置项起始地址冲突：ID：{i - 2}   长度：{val_1} ")
                        FaultFlag = 1
                        break
            else:
                DataBlockNumber = i - 2
                print(f"读取数据块ID结束    总ID数量: {DataBlockNumber}")
                break
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        FaultFlag = 1


def Read_E2DataDefinition(input_file, sheet_name):
    """
    参数:
        input_file: 输入的Excel文件路径
        sheet_name: 要读取的工作表名称
    """
    # 修改全局变量需要声明global
    global Definition_EnumName
    global Definition_Dscribe
    global Definition_StartIndex
    global Definition_Size
    global Definition_BlockID
    global Definition_UseID
    global Definition_UseID_NO_BACK

    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(input_file, data_only=True)

        # 获取指定工作表
        sheet = workbook[sheet_name]

        # 清空全局变量
        Definition_EnumName = []
        Definition_Dscribe = []
        Definition_StartIndex = []
        Definition_Size = []
        Definition_BlockID = []
        Definition_UseID = []
        Definition_UseID_NO_BACK = []

        CustomNumber = 0;  # 数据块ID已经使用的总长度
        ID_Row = 2  # 数据块ID的行数
        CurDataID = sheet.cell(row=ID_Row, column=4).value;  # 取第一个数据块的ID
        UseID_Size = [0] * DataBlockNumber  # 统计使用的数据块ID使用的字节长度
        # 遍历工作表中的所有行
        for i, row in enumerate(sheet.iter_rows(values_only=True), start=2):  # 行号从1开始
            val_1 = sheet.cell(row=i, column=1).value  # 枚举定义
            val_2 = sheet.cell(row=i, column=2).value  # 中文描述
            val_3 = sheet.cell(row=i, column=3).value  # 字节长度
            if val_1 is not None:
                val_1 = str(val_1).strip().replace(' ', '')  # 枚举定义（去除空格）
                # val_2 = str(val_2).strip() if cell_2_val is not None else ''  # 中文描述
                # val_3 = int(str(val_3).strip().replace(' ', ''))  # 字节长度
                Definition_EnumName.append(val_1)
                Definition_Dscribe.append(val_2)
                Definition_Size.append(val_3)
                Definition_BlockID.append(CurDataID)
                Definition_StartIndex.append(CustomNumber)
                CustomNumber += val_3  # 当前数据块长度
            else:
                # 枚举定义为空 数据块定义结束 - 需要判断长度是否小于对应ID
                arrIndex = CurDataID - 1  # 将使用的数据块ID对应记录的数组下标-确保重复的ID可以使用一个数组，后续增加检测重复的ID报错函数
                UseID_Size[arrIndex] += CustomNumber + 1  # 记录当前ID使用的长度
                Definition_UseID.append(CurDataID)
                Definition_UseID_NO_BACK.append(CurDataID)
                if DataBlockAllocation_BackupID[CurDataID] != 0:
                    Definition_UseID.append(DataBlockAllocation_BackupID[CurDataID])
                # print(f"CurDataID = {CurDataID} ")
                if UseID_Size[arrIndex] > DataBlockAllocation_Size[CurDataID]:
                    print(f"错误：数据块{CurDataID}自定义存储项的长度超出范围")
                    print(f"数据块{CurDataID}读取定义结束：使用长度：{UseID_Size[arrIndex]} 总长度：{DataBlockAllocation_Size[CurDataID]}")
                    FaultFlag = 1
                    break
                print(f"数据块{CurDataID}读取定义结束：使用长度：{UseID_Size[arrIndex]} 总长度：{DataBlockAllocation_Size[CurDataID]}")
                CustomNumber = 0
                if val_2 is not None:  # 检测到校验和
                    ID_Row = i + 1
                    CurDataID = sheet.cell(row=ID_Row, column=4).value;  # 取第一个数据块的ID
                    if CurDataID is None:  # 下一个数据块ID为空  检测结束
                        print(f"所有数据块自定义结束")
                        break
    except Exception as e:
        print(f"处理过程中发生错误: {e}")
        FaultFlag = 1


def WriteOutputCFile():
    with open(OutputCFile, 'w') as TxtWrFile:
        LineStr = '/************************************************************************************************\n'
        LineStr = LineStr + StrNowTime + '\n'
        LineStr = LineStr + 'File Name:          ' + OutputCFile + '\n'
        LineStr = LineStr + 'Author:             python Script Generation' + '\n'
        LineStr = LineStr + 'Description:        E2数据块配置定义表 - 使用excel表进行管理，不可手动修改' + '\n'
        LineStr = LineStr + '************************************************************************************************/\n'
        TxtWrFile.write(LineStr)
        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        LineStr = f'#include "{OutputHFile}"\n'
        TxtWrFile.write(LineStr)
        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        for i in range(len(DataBlockAllocation_StarAddr)):
            LineStr = f'INT8U u8MEM_DataBlockID_{i}[{DataBlockAllocation_Size[i]}] = ' + '{0};\n'
            TxtWrFile.write(LineStr)

        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        LineStr = '/* E2数据块配置参数表*/\n'
        LineStr += 'static strEepromDataBlockCfgTable EepromDataBlockCfgTable[] =\n'
        LineStr += '{\n'
        LineStr += '/*  数据块ID  起始地址         长度    	 备份ID    RAM变量地址              			       读写请求状态          			     读写操作结果*/\n'
        TxtWrFile.write(LineStr)

        # for i in range(len(DataBlockAllocation_StarAddr)):
        #     LineStr = '    {'+f'{i},        {DataBlockAllocation_StarAddr[i]},          {DataBlockAllocation_Size[i]},      {DataBlockAllocation_BackupID[i]},      &u8MEM_DataBlockID_{i}[0], EN_EEPROM_STA_IDLE, E2promNullOp'+'},\n'
        #     TxtWrFile.write(LineStr)
        # 计算各字段最大长度- GPT生成
        max_index_len = len(str(len(DataBlockAllocation_StarAddr) - 1))
        max_startaddr_len = max(len(str(addr)) for addr in DataBlockAllocation_StarAddr)
        max_size_len = max(len(str(size)) for size in DataBlockAllocation_Size)
        max_backupid_len = max(len(str(bid)) for bid in DataBlockAllocation_BackupID)
        max_ptr_name_len = max(len(f'&u8MEM_DataBlockID_{i}[0]') for i in range(len(DataBlockAllocation_StarAddr)))

        # 逐行写入内容
        for i in range(len(DataBlockAllocation_StarAddr)):
            index_str = str(i).ljust(max_index_len)
            addr_str = f'0x{DataBlockAllocation_StarAddr[i]:X}'.ljust(5)
            size_str = str(DataBlockAllocation_Size[i]).ljust(max_size_len)
            backupid_str = str(DataBlockAllocation_BackupID[i]).ljust(max_backupid_len)
            ptr_str = f'&u8MEM_DataBlockID_{i}[0]'.ljust(max_ptr_name_len)

            LineStr = f'    {{{index_str},    {addr_str},    {size_str},    {backupid_str},    {ptr_str},    EN_EEPROM_STA_IDLE,    E2promNullOp}},\n'
            TxtWrFile.write(LineStr)

        LineStr = '};\n'
        TxtWrFile.write(LineStr)

        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        LineStr = '/* E2自定义枚举参数表*/\n'
        LineStr += 'static strEepromDataDefineTable EepromDataDefineTable[] =\n'
        LineStr += '{\n'
        LineStr += '    /*枚举定义    												开始下标    字节长度     						数据块ID */\n'
        TxtWrFile.write(LineStr)

        # for i in range(len(Definition_EnumName)):
        #     LineStr = '    {'+f'{Definition_BlockID[i]},                {Definition_StartIndex[i]},      D_{Definition_EnumName[i]}_Len_{Definition_Size[i]},          EN_MemIndex_{Definition_EnumName[i]}'+'},\n'
        #     if i < len(Definition_EnumName )- 1 :
        #         if Definition_BlockID[i] != Definition_BlockID[i+1]:
        #             LineStr += '\n'
        #     TxtWrFile.write(LineStr)
        # LineStr = '};\n'
        # TxtWrFile.write(LineStr)

        # 先预处理每列最大宽度- GPT生成
        max_blockid_len = max(len(str(x)) for x in Definition_BlockID)
        max_startindex_len = max(len(str(x)) for x in Definition_StartIndex)
        max_macro_name_len = max(
            len(f'D_{name}_Len_{size}') for name, size in zip(Definition_EnumName, Definition_Size))
        max_enum_name_len = max(len(f'EN_MemIndex_{name}') for name in Definition_EnumName)

        # 开始写入
        for i in range(len(Definition_EnumName)):
            blockid_str = str(Definition_BlockID[i]).ljust(max_blockid_len)
            startindex_str = str(Definition_StartIndex[i]).ljust(max_startindex_len)
            macro_name_str = f'D_{Definition_EnumName[i]}_Len_{Definition_Size[i]}'.ljust(max_macro_name_len)
            enum_name_str = f'EN_MemIndex_{Definition_EnumName[i]}'.ljust(max_enum_name_len)

            LineStr = f'    {{{enum_name_str},    {startindex_str},    {macro_name_str},    {blockid_str}}},\n'

            if i < len(Definition_EnumName) - 1:
                if Definition_BlockID[i] != Definition_BlockID[i + 1]:
                    LineStr += '\n'

            TxtWrFile.write(LineStr)
        LineStr = '};\n'
        TxtWrFile.write(LineStr)

        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)

        print(f"总共使用ID数量：{len(Definition_UseID)} (包含备份ID)")
        LineStr = '/* 统计使用的ID数量 包含备份ID */\n'
        LineStr += f'INT16U UseDataBlockID[D_USEID_NUMBER]' + ' = {\n'
        for i in range(len(Definition_UseID)):
            LineStr += f'{Definition_UseID[i]}, '
            if i % 10 == 9:
                LineStr += '\n'
        LineStr += '\n'
        LineStr += '};\n'
        TxtWrFile.write(LineStr)
		
        print(f"总共使用ID数量：{len(Definition_UseID_NO_BACK)} (不包含备份ID)")
        LineStr = '/* 统计使用的ID数量 不包含备份ID */\n'
        LineStr += f'INT16U UseDataBlockIDNoBackup[D_USEID_NUMBER_NO_BACKUP]' + ' = {\n'
        for i in range(len(Definition_UseID_NO_BACK)):
            LineStr += f'{Definition_UseID_NO_BACK[i]}, '
            if i % 10 == 9:
                LineStr += '\n'
        LineStr += '\n'
        LineStr += '};\n'
        TxtWrFile.write(LineStr)

        LineStr = '\n\n'
        LineStr += '/**写入常用函数***/\n'
        LineStr = '/***********************************************************************\n'
        LineStr += '* @function name: Eel_GetE2DataBlockCfgTablePtr\n'
        LineStr += '* @description:  获取数据块配置表地址\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'strEepromDataBlockCfgTable * Eel_GetE2DataBlockCfgTablePtr(void)\n'
        LineStr += '{\n'
        LineStr += '    return &EepromDataBlockCfgTable[0];\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockCfgTableLength\n'
        LineStr += '* @description:  获取数据块配置表配置项长度\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetE2DataBlockCfgTableLength(void)\n'
        LineStr += '{\n'
        LineStr += '    return (sizeof(EepromDataBlockCfgTable)/sizeof(EepromDataBlockCfgTable[0]));\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Eel_GetE2DataDefineTablePtr\n'
        LineStr += '* @description:  获取自定义枚举参数表地址\n'
        LineStr += '* @author: python\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'strEepromDataDefineTable * Eel_GetE2DataDefineTablePtr(void)\n'
        LineStr += '{\n'
        LineStr += '    return &EepromDataDefineTable[0];\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataDefineTableLength\n'
        LineStr += '* @description:  获取自定义枚举参数表长度\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetE2DataDefineTableLength(void)\n'
        LineStr += '{\n'
        LineStr += '    return (sizeof(EepromDataDefineTable)/sizeof(EepromDataDefineTable[0]));\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_SetE2DataBlockOptResult\n'
        LineStr += '* @description:  设置数据块ID的操作结果\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'void Srvl_SetE2DataBlockOptResult(INT16U DataId, enE2promOptResult Result)\n'
        LineStr += '{\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        EepromDataBlockCfgTable[DataId].OptResult = Result;\n'
        LineStr += '    }\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockOptResult\n'
        LineStr += '* @description:  获取数据块ID的操作结果\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'enE2promOptResult Srvl_GetE2DataBlockOptResult(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    enE2promOptResult OptResult = E2promNullOp;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        OptResult = EepromDataBlockCfgTable[DataId].OptResult;\n'
        LineStr += '    }\n\n'
        LineStr += '    return OptResult;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_SetE2DataBlockEepOptReq\n'
        LineStr += '* @description:  设置数据块ID的读写请求\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'void Srvl_SetE2DataBlockOptReq(INT16U DataId, EepromOptReq Req)\n'
        LineStr += '{\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        EepromDataBlockCfgTable[DataId].OptReq = Req;\n'
        LineStr += '    }\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockEepOptReq\n'
        LineStr += '* @description:  获取数据块ID的读写请求\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'EepromOptReq Srvl_GetE2DataBlockEepOptReq(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    EepromOptReq OptReq = EN_EEPROM_STA_IDLE;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        OptReq = EepromDataBlockCfgTable[DataId].OptReq;\n'
        LineStr += '    }\n\n'
        LineStr += '    return OptReq;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockBufferPtr\n'
        LineStr += '* @description:  获取数据块ID的RAM变量缓存数据地址\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT8U * Srvl_GetE2DataBlockBufferPtr(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    INT8U * pParaBuffer = NULL;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        pParaBuffer = EepromDataBlockCfgTable[DataId].pParaBuffer;\n'
        LineStr += '    }\n\n'
        LineStr += '    return pParaBuffer;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockLength\n'
        LineStr += '* @description:  获取数据块ID的RAM变量配置的长度\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT8U Srvl_GetE2DataBlockLength(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    INT8U DataLength = 0;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        DataLength = EepromDataBlockCfgTable[DataId].DataLength;\n'
        LineStr += '    }\n\n'
        LineStr += '    return DataLength;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockStartAddr\n'
        LineStr += '* @description:  获取数据块ID的起始地址\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT32U Srvl_GetE2DataBlockStartAddr(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    INT32U StartAddr = 0;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        StartAddr = EepromDataBlockCfgTable[DataId].StartAddr;\n'
        LineStr += '    }\n\n'
        LineStr += '    return StartAddr;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DataBlockBackUpID\n'
        LineStr += '* @description:  获取数据块ID的备份ID\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetE2DataBlockBackUpID(INT16U DataId)\n'
        LineStr += '{\n'
        LineStr += '    INT16U BackUpID = 0;\n\n'
        LineStr += '    if(DataId < Srvl_GetE2DataBlockCfgTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        BackUpID = EepromDataBlockCfgTable[DataId].BackUpID;\n'
        LineStr += '    }\n\n'
        LineStr += '    return BackUpID;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DefineDataBlockID\n'
        LineStr += '* @description:  获取E2定义数据块ID\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetE2DefineDataBlockID(enSrvMemIndex MemIndex)\n'
        LineStr += '{\n'
        LineStr += '    INT16U DataBlockID = 0;\n\n'
        LineStr += '    if(MemIndex < Srvl_GetE2DataDefineTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        DataBlockID = EepromDataDefineTable[MemIndex].DataBlockID;\n'
        LineStr += '    }\n\n'
        LineStr += '    return DataBlockID;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DefineLen\n'
        LineStr += '* @description:  获取E2定义长度\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT8U Srvl_GetE2DefineLen(enSrvMemIndex MemIndex)\n'
        LineStr += '{\n'
        LineStr += '    INT8U len = 0;\n\n'
        LineStr += '    if(MemIndex < Srvl_GetE2DataDefineTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        len = EepromDataDefineTable[MemIndex].len;\n'
        LineStr += '    }\n\n'
        LineStr += '    return len;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetE2DefineStartOffset\n'
        LineStr += '* @description:  获取E2定义起始偏移\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT8U Srvl_GetE2DefineStartOffset(enSrvMemIndex MemIndex)\n'
        LineStr += '{\n'
        LineStr += '    INT8U StartOffset = 0;\n\n'
        LineStr += '    if(MemIndex < Srvl_GetE2DataDefineTableLength())\n'
        LineStr += '    {\n'
        LineStr += '        StartOffset = EepromDataDefineTable[MemIndex].StartOffset;\n'
        LineStr += '    }\n\n'
        LineStr += '    return StartOffset;\n'
        LineStr += '}\n\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetUseDataBlockID\n'
        LineStr += '* @description:  获取使用的数据块ID\n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetUseDataBlockID(INT16U Index)\n'
        LineStr += '{\n'
        LineStr += '    INT16U DataBlockID = 0;\n\n'
        LineStr += '    if(Index < D_USEID_NUMBER)\n'
        LineStr += '    {\n'
        LineStr += '        DataBlockID = UseDataBlockID[Index];\n'
        LineStr += '    }\n\n'
        LineStr += '    return DataBlockID;\n'
        LineStr += '}\n'

        LineStr += '/***********************************************************************\n'
        LineStr += '* @function name: Srvl_GetUseDataBlockIDNoBackUp\n'
        LineStr += '* @description:  获取使用的数据块ID 不包含备份ID \n'
        LineStr += '* @author: python Script Generation\n'
        LineStr += f'* @{StrNowTime}\n'
        LineStr += '***********************************************************************/\n'
        LineStr += 'INT16U Srvl_GetUseDataBlockIDNoBackUp(INT16U Index)\n'
        LineStr += '{\n'
        LineStr += '    INT16U DataBlockID = 0;\n\n'
        LineStr += '    if(Index < D_USEID_NUMBER_NO_BACKUP)\n'
        LineStr += '    {\n'
        LineStr += '        DataBlockID = UseDataBlockIDNoBackup[Index];\n'
        LineStr += '    }\n\n'
        LineStr += '    return DataBlockID;\n'
        LineStr += '}\n'

        TxtWrFile.write(LineStr)
    print(f"写入{OutputCFile}  ok")


def WriteOutputHFile():
    with open(OutputHFile, 'w') as TxtWrFile:
        LineStr = '/************************************************************************************************\n'
        LineStr = LineStr + StrNowTime + '\n'
        LineStr = LineStr + 'File Name:          ' + OutputHFile + '\n'
        LineStr = LineStr + 'Author:             python Script Generation' + '\n'
        LineStr = LineStr + 'Description:        E2数据块配置定义表 - 使用excel表进行管理，不可手动修改' + '\n'
        LineStr = LineStr + '************************************************************************************************/\n'
        TxtWrFile.write(LineStr)
        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        LineStr = '#ifndef _PYTHONE_E2CFG_H_' + '\n'
        LineStr = LineStr + '#define _PYTHONE_E2CFG_H_' + '\n\n'
        LineStr = LineStr + '#include "Common.h"' + '\n'
        TxtWrFile.write(LineStr)
        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)

        LineStr = '/**数据块读写请求枚举**/\n';
        LineStr += 'typedef enum _EEPROM_OPT_STA_T\n';
        LineStr += '{\n';
        LineStr += '    EN_EEPROM_STA_IDLE = 0, /*空闲*/\n';
        LineStr += '    EN_EEPROM_STA_WRITE,    /*写请求状态*/\n';
        LineStr += '    EN_EEPROM_STA_READ,     /*读请求状态*/\n';
        LineStr += '} EepromOptReq;\n\n';

        LineStr += '/**数据块读写结果枚举定义**/\n';
        LineStr += 'typedef enum\n';
        LineStr += '{\n';
        LineStr += '    E2promNullOp  = 0,   /*初始化*/\n';
        LineStr += '    E2promWaitIDL,       /*等待操作*/\n';
        LineStr += '    E2promOping,         /*操作中*/\n';
        LineStr += '    E2promOpOK,          /*操作成功*/\n';
        LineStr += '    E2promOpFail,        /*操作失败*/\n';
        LineStr += '} enE2promOptResult;\n\n';

        LineStr += '/**数据块配置数据结构**/\n';
        LineStr += 'typedef struct EEPROMPARATABLE\n';
        LineStr += '{\n';
        LineStr += '    INT16U             DataBlockID;         /*ID标识*/\n';
        LineStr += '    INT32U             StartAddr;           /*起始地址*/\n';
        LineStr += '    INT8U              DataLength;          /*数据长度*/\n';
        LineStr += '    INT16U             BackUpID;            /*备份数据标识*/\n';
        LineStr += '    INT8U              *pParaBuffer;        /*指向的数据*/\n';
        LineStr += '    EepromOptReq       OptReq;              /*读写请求状态*/\n';
        LineStr += '    enE2promOptResult  OptResult;           /*读写操作结果*/\n';
        LineStr += '} strEepromDataBlockCfgTable;\n\n';

        LineStr += '/**自定义存储项数据结构**/\n';
        LineStr += 'typedef struct EE_SRV_DATA_TABLE_CFG\n';
        LineStr += '{\n';
        LineStr += '    INT16U             SrvMemIndex;     /* 自定义枚举 */\n';
        LineStr += '    INT8U              StartOffset;     /* 定义存储在数据块中对应RAM变量的起始偏移量 */\n';
        LineStr += '    INT8U              len;             /* 定义存储长度 */\n';
        LineStr += '    INT16U             DataBlockID;     /* 数据块ID */\n';
        LineStr += '} strEepromDataDefineTable;\n';

        TxtWrFile.write(LineStr)

        LineStr = '\n\n'
        TxtWrFile.write(LineStr)

        LineStr = f'#define    D_USEID_NUMBER          {len(Definition_UseID)}     /* 使用ID的数量 */\n'
        LineStr += f'#define    D_USEID_NUMBER_NO_BACKUP      {len(Definition_UseID_NO_BACK)}  /* 使用ID的数量 不包含备份*/\n'
        LineStr += f'#define    D_NOBACKUP              0     /* 无备份*/\n'
        TxtWrFile.write(LineStr)
        LineStr = '\n\n'
        TxtWrFile.write(LineStr)
        # 宏定义存储项长度
        # for i in range(len(Definition_EnumName)):
        #     LineStr = f'#define     D_{Definition_EnumName[i]}_Len_{Definition_Size[i]}        {Definition_Size[i]}\n'
        #     if i < (len(Definition_EnumName) - 1):
        #         if Definition_BlockID[i] != Definition_BlockID[i+1]:
        #             LineStr += '\n'
        #     TxtWrFile.write(LineStr)
        # 计算最长的宏名长度，用于对齐 - GPT生成
        max_macro_name_len = max(
            len(f'D_{name}_Len_{size}') for name, size in zip(Definition_EnumName, Definition_Size))

        for i in range(len(Definition_EnumName)):
            macro_name = f'D_{Definition_EnumName[i]}_Len_{Definition_Size[i]}'
            macro_value = Definition_Size[i]

            # 格式化对齐，确保等号和数值对齐
            LineStr = f'#define     {macro_name.ljust(max_macro_name_len)}    {macro_value}\n'

            if i < (len(Definition_EnumName) - 1):
                if Definition_BlockID[i] != Definition_BlockID[i + 1]:
                    LineStr += '\n'

            TxtWrFile.write(LineStr)

        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)

        # 枚举定义存储项序号 - 第一个手动赋值0
        LineStr = 'typedef enum\n'
        LineStr += '{\n'
        LineStr += f'/********数据块ID_{Definition_BlockID[0]} 定义开始********/\n'
        LineStr += '    #if 0\n'
        LineStr += f'    D:{Definition_Dscribe[0]} \n'
        LineStr += '    #endif\n'
        LineStr += f'    EN_MemIndex_{Definition_EnumName[0]} = 0, /* D_{Definition_EnumName[0]}_Len_{Definition_Size[0]} */\n'
        if Definition_BlockID[0] != Definition_BlockID[1]:
            LineStr += f'/********数据块ID_{Definition_BlockID[0]} 定义结束********/\n'
            LineStr += '\n'
            LineStr += f'/********数据块ID_{Definition_BlockID[1]} 定义开始********/\n'
        TxtWrFile.write(LineStr)
        
        for i in range(1, len(Definition_EnumName)):
            # LineStr = '    #if 0\n'
            # LineStr += f'中文描述 : {Definition_Dscribe[i]}\n'
            # LineStr += '    #endif\n'
            # LineStr += f'    EN_MemIndex_{Definition_EnumName[i]}, /* D_{Definition_EnumName[i]}_Len_{Definition_Size[i]} */\n'
            
            # 调整注释缩进
            desc_lines = Definition_Dscribe[i].splitlines()
            desc_with_tabs = '\n'.join(['\t' + line for line in desc_lines])
            
            LineStr = '    #if 0\n'
            LineStr += f'{desc_with_tabs}\n'
            LineStr += '    #endif\n'
            LineStr += f'    EN_MemIndex_{Definition_EnumName[i]}, /* D_{Definition_EnumName[i]}_Len_{Definition_Size[i]} */\n'

            if i < len(Definition_EnumName) - 1:
                if Definition_BlockID[i] != Definition_BlockID[i + 1]:
                    LineStr += f'/********数据块ID_{Definition_BlockID[i]} 定义结束********/\n'
                    LineStr += '\n'
                    LineStr += f'/********数据块ID_{Definition_BlockID[i + 1]} 定义开始********/\n'
            else:
                LineStr += f'/********数据块ID_{Definition_BlockID[i]} 定义结束********/\n'
                LineStr += f'    EN_MemIndex_Max\n'
            TxtWrFile.write(LineStr)

        LineStr = '} enSrvMemIndex;\n'
        TxtWrFile.write(LineStr)

        LineStr = '\n\n\n'
        TxtWrFile.write(LineStr)
        #
        LineStr = '/***写入常用函数声明***/\n';
        LineStr += 'extern strEepromDataBlockCfgTable * Eel_GetE2DataBlockCfgTablePtr(void);\n';
        LineStr += 'extern INT16U Srvl_GetE2DataBlockCfgTableLength(void);\n';
        LineStr += 'extern strEepromDataDefineTable * Eel_GetE2DataDefineTablePtr(void);\n';
        LineStr += 'extern INT16U Srvl_GetE2DataDefineTableLength(void);\n';
        LineStr += 'extern void Srvl_SetE2DataBlockOptResult(INT16U DataId, enE2promOptResult Result);\n';
        LineStr += 'extern enE2promOptResult Srvl_GetE2DataBlockOptResult(INT16U DataId);\n';
        LineStr += 'extern void Srvl_SetE2DataBlockOptReq(INT16U DataId, EepromOptReq Req);\n';
        LineStr += 'extern EepromOptReq Srvl_GetE2DataBlockEepOptReq(INT16U DataId);\n';
        LineStr += 'extern INT8U * Srvl_GetE2DataBlockBufferPtr(INT16U DataId);\n';
        LineStr += 'extern INT8U Srvl_GetE2DataBlockLength(INT16U DataId);\n';
        LineStr += 'extern INT32U Srvl_GetE2DataBlockStartAddr(INT16U DataId);\n';
        LineStr += 'extern INT16U Srvl_GetE2DataBlockBackUpID(INT16U DataId);\n';
        LineStr += 'extern INT16U Srvl_GetE2DefineDataBlockID(enSrvMemIndex MemIndex);\n';
        LineStr += 'extern INT8U Srvl_GetE2DefineLen(enSrvMemIndex MemIndex);\n';
        LineStr += 'extern INT8U Srvl_GetE2DefineStartOffset(enSrvMemIndex MemIndex);\n';
        LineStr += 'extern INT16U Srvl_GetUseDataBlockID(INT16U Index);\n';
        LineStr += 'extern INT16U Srvl_GetUseDataBlockIDNoBackUp(INT16U Index);\n';
        TxtWrFile.write(LineStr)

        LineStr = '\n'
        LineStr += '#endif\n'
        TxtWrFile.write(LineStr)

    print(f"写入{OutputHFile}  ok")


def print_E2DataBlockAllocation_data():
    """
    打印全局变量中的数据
    """
    # global DataBlockAllocation_StarAddr

    if not DataBlockAllocation_StarAddr:
        print("没有有效数据可打印")
        return

    print("\n打印有效数据:")

    print(f"ID   起始地址  长度 备份ID")
    for i, row in enumerate(DataBlockAllocation_StarAddr, 0):
        if ((DataBlockAllocation_StarAddr[i] + DataBlockAllocation_Size[i]) <= DataBlockAllocation_StarAddr[
            i + 1]) and i < len(DataBlockAllocation_StarAddr):
            print(
                f" {i + 1}   {DataBlockAllocation_StarAddr[i]}      {DataBlockAllocation_Size[i]}  {DataBlockAllocation_BackupID[i]} ")
        else:
            print(f" ID {i + 1} 长度设置错误:")
            break


def print_E2DataDefinition_data():
    """
    打印全局变量中的数据
    """
    # global DataBlockAllocation_StarAddr

    if not Definition_Dscribe:
        print("没有有效数据可打印")
        return

    print("\n打印有效数据:")

    # print(f"枚举定义 - 开始下标 - 长度 - E2定义项 ")
    # for i, row in enumerate(Definition_EnumName, 0):
    #     print(f"{Definition_EnumName[i]}   {Definition_StartIndex[i]}      {Definition_Size[i]}   {Definition_Dscribe[i]}   ")
    # print(f"统计使用的ID ")
    # for i, row in enumerate(Definition_UseID, 0):
    #     print(f"{Definition_UseID[i]} ")


# 使用示例
if __name__ == "__main__":

    # 1. 获取数据
    Read_E2DataBlockAllocation(input_excel, sheet_name0)
    if FaultFlag == 0:
        Read_E2DataDefinition(input_excel, sheet_name1)

    # 2. 打印数据
    # print_E2DataBlockAllocation_data()
    # print_E2DataDefinition_data()

    # 2. 读取文件无错误直接生成。c.h文件
    if FaultFlag == 0:
        WriteOutputCFile()
        WriteOutputHFile()
    print("程序已经执行完成！")
    # inp = input("请按回车键退出程序。")
